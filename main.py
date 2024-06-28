import concurrent.futures
import tkinter as tk
from tkinter import messagebox
from tkinter import filedialog
from datetime import datetime
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import webbrowser
from tkinter import ttk
import threading

# Variables globales
proceso_en_ejecucion = False
estado_codigos = []
total_codigos = 0
codigos_procesados = 0

# Función para obtener el estado de un producto
def obtener_estado(codigo_padre, pais):
    url_base = f'https://www.marathon.store/{pais}/view/ProductVariantSelectorComponentController?componentUid=VariantSelector&currentProductCode='
    url = url_base + str(codigo_padre)
    response = requests.get(url)
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')
        lis = soup.find_all('li', attrs={'data-url': lambda x: x and str(codigo_padre) in x})
        if lis:
            for li in lis:
                data_has_stock = li.get('data-has-stock')
                data_selected = li.get('data-selected')
                if data_has_stock == "false":
                    return "Agotado"
                elif data_has_stock == "true":
                    return "Disponible"
                elif data_selected == "true":
                    return "Disponible"
        else:
            return "Agotado"
    elif response.status_code == 404:
        return "ERROR 404"
    else:
        return "Error de conexión"

# Función para procesar los códigos y guardar en Excel
def procesar_codigos(codigos, pais):
    global proceso_en_ejecucion, estado_codigos, total_codigos, codigos_procesados
    total_codigos = len(codigos)
    start_time = datetime.now()
    
    def obtener_estado_concurrente(codigo_padre, pais):
        url_base = f'https://www.marathon.store/{pais}/view/ProductVariantSelectorComponentController?componentUid=VariantSelector&currentProductCode='
        url = url_base + str(codigo_padre)
        response = requests.get(url)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'html.parser')
            lis = soup.find_all('li', attrs={'data-url': lambda x: x and str(codigo_padre) in x})
            if lis:
                for li in lis:
                    data_has_stock = li.get('data-has-stock')
                    data_selected = li.get('data-selected')
                    if data_has_stock == "false":
                        return "Agotado"
                    elif data_has_stock == "true":
                        return "Disponible"
                    elif data_selected == "true":
                        return "Disponible"
            else:
                return "Agotado"
        elif response.status_code == 404:
            return "ERROR 404"
        else:
            return "Error de conexión"
    
    with concurrent.futures.ThreadPoolExecutor() as executor:
        futures = {executor.submit(obtener_estado_concurrente, codigo, pais): codigo for codigo in codigos}
        for future in concurrent.futures.as_completed(futures):
            codigo = futures[future]
            try:
                estado = future.result()
            except Exception as exc:
                estado = f"Error: {exc}"
            estado_codigos.append((codigo, estado))
            codigos_procesados += 1
            # Actualizar información en la interfaz
            elapsed_time = datetime.now() - start_time
            tiempo_transcurrido = str(elapsed_time).split('.')[0]  # Formato HH:MM:SS
            info_estado.set(f"Procesando código {codigos_procesados}/{total_codigos} - Tiempo transcurrido: {tiempo_transcurrido}")
            barra_progreso['value'] = (codigos_procesados / total_codigos) * 100
            root.update_idletasks()

    if proceso_en_ejecucion:
        guardar_resultados(pais)

# Función para pausar el proceso
def pausar_proceso():
    global proceso_en_ejecucion
    proceso_en_ejecucion = False
    messagebox.showinfo("Proceso pausado", "Se ha pausado el proceso. Puede continuar luego.")

# Función para detener el proceso
def detener_proceso():
    global proceso_en_ejecucion
    proceso_en_ejecucion = False
    guardar_resultados(pais_seleccionado.get())
    messagebox.showinfo("Proceso detenido", "Se ha detenido el proceso.")
    root.quit()

# Función para guardar los resultados en Excel
def guardar_resultados(pais):
    global proceso_en_ejecucion
    # Guardar en Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Control Stock Web"
    ws['A1'] = "CODIGO"
    ws['B1'] = "STATUS WEB"
    for i, (codigo, estado) in enumerate(estado_codigos, start=2):
        ws[f'A{i}'] = codigo
        ws[f'B{i}'] = estado

    # Guardar archivo
    fecha_actual = datetime.now().strftime("%Y-%m-%d")
    nombre_archivo = f"Control Stock Web {pais.upper()} {fecha_actual}.xlsx"
    wb.save(nombre_archivo)
    messagebox.showinfo("Proceso completado", f"Se ha guardado el archivo '{nombre_archivo}' con los estados de los productos.")
    # Abrir archivo Excel
    webbrowser.open(nombre_archivo)

# Función para iniciar el procesamiento en un hilo separado
def iniciar_procesamiento():
    global proceso_en_ejecucion
    if not proceso_en_ejecucion:
        proceso_en_ejecucion = True
        pais = pais_seleccionado.get()
        codigos = entry_codigos.get("1.0", "end").split()
        threading.Thread(target=procesar_codigos, args=(codigos, pais)).start()

# Interfaz gráfica
root = tk.Tk()
root.title("Verificar Stock Web")

# Entrada de códigos
tk.Label(root, text="Ingrese los códigos separados por espacio:").pack()
entry_codigos = tk.Text(root, height=10, width=50)
entry_codigos.pack()

# Botones de selección de país
tk.Label(root, text="Seleccione un país:").pack()
paises = [("Perú", "pe"), ("Ecuador", "ec"), ("Bolivia", "bo"), ("Chile", "ch")]
pais_seleccionado = tk.StringVar()
for nombre_pais, codigo_pais in paises:
    tk.Radiobutton(root, text=nombre_pais, variable=pais_seleccionado, value=codigo_pais).pack()

# Botón de iniciar procesamiento
tk.Button(root, text="Iniciar Procesamiento", command=iniciar_procesamiento).pack()

# Botón de detener
btn_detener = tk.Button(root, text="Detener", command=detener_proceso)
btn_detener.pack()

# Información de estado y progreso
info_estado = tk.StringVar()
tk.Label(root, textvariable=info_estado).pack()

barra_progreso = ttk.Progressbar(root, length=200, mode='determinate')
barra_progreso.pack()

root.mainloop()
